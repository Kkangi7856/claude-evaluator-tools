# -*- coding: utf-8 -*-
"""
render_summary.py — 평가위원 답변 엑셀 → 항목별 「평가의견 종합」 .hwp 생성

두 가지 기능:
  1) --extract : 엑셀에서 평가위원별 원문 의견을 JSON 으로 추출(중복 제거는 Claude 가 수행).
  2) --spec    : 항목별로 정리된 스펙(JSON)을 받아 네이티브 .hwp 생성.

사용:
  # 원문 추출
  python render_summary.py --extract "평가표 답변 목록.xlsx"
  # 문서 생성
  python render_summary.py --spec summary_spec.json --outdir .

스펙 예시는 references/sample-spec.json 참조.
"""
import argparse, json, os, re, sys

CATEGORIES = ("우수한 점", "수정 및 보완할 점", "기타 의견")


# ---------- 유틸 ----------
def col_to_idx(col):
    """'A'->1, 'C'->3 (엑셀 열 문자 → 1기반 인덱스). 숫자면 그대로 int."""
    col = str(col).strip()
    if col.isdigit():
        return int(col)
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def clean_item(text):
    """항목 문두의 기호/번호(ㅇ, ○, -, ·, 1), 1. 등)를 제거해 정규화."""
    t = (text or "").strip()
    t = re.sub(r"^\s*(?:[ㅇ○◦·•\-*]|\d+[.)]|[가-힣][.)])\s*", "", t)
    return t.strip()


# ---------- 엑셀 원문 추출 ----------
def extract(args):
    import openpyxl
    wb = openpyxl.load_workbook(args.extract, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]
    ncol = col_to_idx(args.name_col)
    gcol = col_to_idx(args.good_col)
    fcol = col_to_idx(args.fix_col)
    ecol = col_to_idx(args.etc_col)
    rows = []
    for r in range(args.start_row, ws.max_row + 1):
        name = ws.cell(r, ncol).value
        if name is None or not str(name).strip():
            continue
        rows.append({
            "name": str(name).strip(),
            CATEGORIES[0]: (ws.cell(r, gcol).value or ""),
            CATEGORIES[1]: (ws.cell(r, fcol).value or ""),
            CATEGORIES[2]: (ws.cell(r, ecol).value or ""),
        })
    print(json.dumps(rows, ensure_ascii=False, indent=2))
    return 0


# ---------- .hwp 생성 ----------
def write_para(hwp, text="", bold=False, size=11, align="left", indent=0, hanging=False):
    if align == "center":
        hwp.ParagraphShapeAlignCenter()
    else:
        hwp.ParagraphShapeAlignLeft()
    hwp.set_font(Bold=bold, Height=size)
    if indent:
        text = (" " * indent) + text
    hwp.insert_text(text)
    hwp.BreakPara()


def build_summary(hwp, spec, outpath):
    title = spec.get("title", "평가의견 종합")
    hwp.set_font(FaceName="함초롬바탕", Height=11)

    write_para(hwp, title, bold=True, size=16, align="center")
    write_para(hwp, "", size=11)

    for sec in spec.get("sections", []):
        heading = str(sec.get("heading", "")).strip()
        items = sec.get("items") or []
        if not heading and not items:
            continue
        write_para(hwp, heading, bold=True, size=13)
        if items:
            for it in items:
                t = clean_item(str(it))
                if not t:
                    continue
                write_para(hwp, f"ㅇ {t}", size=11, indent=2)
        else:
            write_para(hwp, "ㅇ 해당 의견 없음", size=11, indent=2)
        write_para(hwp, "", size=11)

    hwp.MoveDocEnd()
    hwp.save_as(outpath, "HWP")
    try:
        hwp.save_as(outpath[:-4] + ".pdf", "PDF")
    except Exception:
        pass


def safe_name(s):
    for ch in '\\/:*?"<>|':
        s = s.replace(ch, "")
    return s.strip()


def render(args):
    from pyhwpx import Hwp
    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    title = safe_name(spec.get("title", "평가의견 종합"))
    out_name = spec.get("out_name") or f"{title}.hwp"
    if not out_name.lower().endswith(".hwp"):
        out_name += ".hwp"
    os.makedirs(args.outdir, exist_ok=True)
    outpath = os.path.join(args.outdir, out_name)

    hwp = Hwp(visible=False, new=True)
    try:
        build_summary(hwp, spec, outpath)
    finally:
        hwp.quit()
    print(f"[완료] 평가의견 종합: {outpath}")
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser(description="평가위원 답변 엑셀 → 항목별 「평가의견 종합」 .hwp 생성")
    ap.add_argument("--extract", help="원문 추출할 엑셀 파일 경로")
    ap.add_argument("--spec", help="항목별 정리 스펙 JSON 파일")
    ap.add_argument("--outdir", default=".", help="출력 폴더")
    # 추출 옵션(열 위치)
    ap.add_argument("--sheet", default=None, help="시트명(생략 시 첫 시트)")
    ap.add_argument("--start-row", type=int, default=2, help="데이터 시작 행(기본 2)")
    ap.add_argument("--name-col", default="A", help="평가위원명 열(기본 A)")
    ap.add_argument("--good-col", default="C", help="우수한 점 열(기본 C)")
    ap.add_argument("--fix-col", default="D", help="수정 및 보완할 점 열(기본 D)")
    ap.add_argument("--etc-col", default="E", help="기타 의견 열(기본 E)")
    a = ap.parse_args(argv)

    if a.extract:
        return extract(a)
    if a.spec:
        return render(a)
    ap.error("--extract 또는 --spec 중 하나는 필요합니다.")


if __name__ == "__main__":
    sys.exit(main())
