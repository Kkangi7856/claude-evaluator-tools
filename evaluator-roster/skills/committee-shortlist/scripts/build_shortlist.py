# -*- coding: utf-8 -*-
"""
build_shortlist.py

선정평가위원 및 후보자 명단 + 단계평가위원 및 후보자 명단 → 최종평가위원 섭외우선(안)

- 소스 2개 파일에서 C~K열(성명·소속·부서·직위·국가연구자번호·전공·기술분류1~3)을 가져오고,
- W열(참석여부)이 'Y'면 실제 참여 '위원', 공란이면 '후보자'로 판정하여
- B열(기평가 섭외이력)을 산출한다.
- K열(국가연구자번호)을 고유키로 두 파일을 병합·중복 제거하여 1인 1행으로 출력한다.

규칙(확정):
  선정=위원, 단계=없음            -> "선정평가위원"
  선정=없음, 단계=위원            -> "단계평가위원"
  선정=후보자, 단계=없음          -> "선정평가위원 후보자"
  선정=없음, 단계=후보자          -> "단계평가위원 후보자"
  선정=후보자, 단계=후보자        -> "선정/단계평가위원 후보자"
  선정=위원,   단계=위원          -> "선정/단계평가위원"
  선정=위원,   단계=후보자        -> "선정평가위원"   (위원 우선)
  선정=후보자, 단계=위원          -> "단계평가위원"   (위원 우선)

정렬(섭외우선): 위원 -> 후보자 순, 같은 그룹 내에선 원본 등장 순서 유지.
"""

import argparse
import sys
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

# ---- 소스 파일 열 위치(1-indexed) ----
SRC = {
    "name": 10,   # J 후보자 성명
    "num": 11,    # K 국가연구자번호 (고유키)
    "org": 12,    # L 소속기관
    "dept": 13,   # M 부서
    "title": 14,  # N 직위
    "major": 18,  # R 전공
    "tc1": 19,    # S 기술분류1
    "tc2": 20,    # T 기술분류2
    "tc3": 21,    # U 기술분류3
    "attend": 23, # W 참석여부
}
SRC_FIRST_DATA_ROW = 3  # 1: 제목, 2: 헤더, 3~: 데이터

# ---- 대상 파일 열 위치(1-indexed) ----
TGT = {"gubun": 1, "history": 2, "name": 3, "org": 4, "dept": 5,
       "title": 6, "num": 7, "major": 8, "tc1": 9, "tc2": 10, "tc3": 11}
TGT_HEADER_ROW = 6
TGT_FIRST_DATA_ROW = 7

# ---- B열 우선순위(정렬용) ----
PRIORITY = {
    "선정/단계평가위원": 0,
    "선정평가위원": 1,
    "단계평가위원": 2,
    "선정/단계평가위원 후보자": 3,
    "선정평가위원 후보자": 4,
    "단계평가위원 후보자": 5,
}


def norm(v):
    return "" if v is None else str(v).strip()


def is_attended(w):
    return norm(w).upper() == "Y"


def scan_source(path):
    """파일을 읽어 {국가연구자번호: {status, data, order}} 반환.
    status: 'member'(위원) | 'candidate'(후보자)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    people = {}
    order = 0
    for r in range(SRC_FIRST_DATA_ROW, ws.max_row + 1):
        num = ws.cell(row=r, column=SRC["num"]).value
        name = ws.cell(row=r, column=SRC["name"]).value
        if num is None and (name is None or norm(name) == ""):
            continue  # 빈 행
        key = num if num is not None else f"__noid__{name}_{r}"
        attended = is_attended(ws.cell(row=r, column=SRC["attend"]).value)
        data = {f: ws.cell(row=r, column=SRC[f]).value for f in
                ("name", "num", "org", "dept", "title", "major", "tc1", "tc2", "tc3")}
        if key not in people:
            people[key] = {"status": "candidate", "data": data, "order": order}
            order += 1
        # 같은 파일 내 같은 번호가 여러 번이면 하나라도 참석 Y이면 위원
        if attended:
            people[key]["status"] = "member"
            people[key]["data"] = data  # 위원 행 정보를 우선 채택
    return people


def decide_history(sel, stg):
    """sel/stg: None | 'member' | 'candidate' -> B열 문자열"""
    if sel == "member" and stg == "member":
        return "선정/단계평가위원"
    if sel == "candidate" and stg == "candidate":
        return "선정/단계평가위원 후보자"
    # 혼합 케이스: 위원 우선
    if sel == "member":            # 단계는 없음 또는 후보자
        return "선정평가위원"
    if stg == "member":            # 선정은 없음 또는 후보자
        return "단계평가위원"
    if sel == "candidate":
        return "선정평가위원 후보자"
    if stg == "candidate":
        return "단계평가위원 후보자"
    return ""  # 도달 불가


def pick_data(sel_rec, stg_rec, sel_status, stg_status):
    """C~K에 채울 원본 데이터 선택. 위원 행 우선, 그다음 선정 파일 우선."""
    order = []
    if sel_status == "member":
        order.append(sel_rec)
    if stg_status == "member":
        order.append(stg_rec)
    if sel_rec:
        order.append(sel_rec)
    if stg_rec:
        order.append(stg_rec)
    for rec in order:
        if rec:
            return rec["data"]
    return {}


def merge(sel_people, stg_people):
    keys = list(dict.fromkeys(list(sel_people) + list(stg_people)))  # 등장 순서 보존
    rows = []
    for k in keys:
        sel_rec = sel_people.get(k)
        stg_rec = stg_people.get(k)
        sel_status = sel_rec["status"] if sel_rec else None
        stg_status = stg_rec["status"] if stg_rec else None
        history = decide_history(sel_status, stg_status)
        data = pick_data(sel_rec, stg_rec, sel_status, stg_status)
        # 원본 첫 등장 순서(선정 우선): 안정 정렬용 보조키
        base_order = sel_rec["order"] if sel_rec else (10_000 + stg_rec["order"])
        rows.append({"history": history, "data": data, "order": base_order})
    rows.sort(key=lambda x: (PRIORITY.get(x["history"], 99), x["order"]))
    return rows


def unmerge_below_header(ws, first_row):
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= first_row:
            ws.unmerge_cells(str(rng))


def build(seon_path, dan_path, template_path, out_path, panel=None):
    sel_people = scan_source(seon_path)
    stg_people = scan_source(dan_path)
    rows = merge(sel_people, stg_people)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 기존 안내문(※...) 텍스트 확보
    note_text = None
    for r in range(TGT_FIRST_DATA_ROW, ws.max_row + 1):
        v = ws.cell(row=r, column=TGT["gubun"]).value
        if isinstance(v, str) and v.strip().startswith("※"):
            note_text = v
            break

    # A열 구분(패널) 기본값: 템플릿 기존 A7 값 유지
    if panel is None:
        panel = ws.cell(row=TGT_FIRST_DATA_ROW, column=TGT["gubun"]).value

    # 데이터 영역 병합 해제 후 기존 내용/스타일 원본 확보
    style_src = {c: copy(ws.cell(row=TGT_FIRST_DATA_ROW, column=c)._style)
                 for c in range(1, len(TGT) + 1)}
    unmerge_below_header(ws, TGT_FIRST_DATA_ROW)

    # 기존 데이터 영역 내용 삭제
    for r in range(TGT_FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    # 데이터 기록
    field_to_col = [("history", "history"), ("name", "name"), ("org", "org"),
                    ("dept", "dept"), ("title", "title"), ("num", "num"),
                    ("major", "major"), ("tc1", "tc1"), ("tc2", "tc2"), ("tc3", "tc3")]
    for i, row in enumerate(rows):
        r = TGT_FIRST_DATA_ROW + i
        # 스타일 복사(테두리/정렬 유지)
        for c in range(1, len(TGT) + 1):
            ws.cell(row=r, column=c)._style = copy(style_src[c])
        ws.cell(row=r, column=TGT["history"]).value = row["history"]
        d = row["data"]
        ws.cell(row=r, column=TGT["name"]).value = d.get("name")
        ws.cell(row=r, column=TGT["org"]).value = d.get("org")
        ws.cell(row=r, column=TGT["dept"]).value = d.get("dept")
        ws.cell(row=r, column=TGT["title"]).value = d.get("title")
        ws.cell(row=r, column=TGT["num"]).value = d.get("num")
        ws.cell(row=r, column=TGT["major"]).value = d.get("major")
        ws.cell(row=r, column=TGT["tc1"]).value = d.get("tc1")
        ws.cell(row=r, column=TGT["tc2"]).value = d.get("tc2")
        ws.cell(row=r, column=TGT["tc3"]).value = d.get("tc3")

    n = len(rows)
    last_data_row = TGT_FIRST_DATA_ROW + n - 1

    # A열(구분) 데이터 행 병합 + 패널명
    if n > 0:
        ws.merge_cells(start_row=TGT_FIRST_DATA_ROW, start_column=TGT["gubun"],
                       end_row=last_data_row, end_column=TGT["gubun"])
        ws.cell(row=TGT_FIRST_DATA_ROW, column=TGT["gubun"]).value = panel

    # 안내문을 데이터 아래로 이동
    if note_text:
        ws.cell(row=last_data_row + 2, column=TGT["gubun"]).value = note_text

    wb.save(out_path)
    return rows, n


def main(argv=None):
    p = argparse.ArgumentParser(
        description="선정/단계 평가위원 명단을 최종 섭외우선(안)으로 병합")
    p.add_argument("--seon", default="선정평가위원 및 후보자 명단.xlsx",
                   help="선정평가위원 및 후보자 명단 파일")
    p.add_argument("--dan", default="단계평가위원 및 후보자 명단.xlsx",
                   help="단계평가위원 및 후보자 명단 파일")
    p.add_argument("--template", default="최종평가위원 섭외우선(안).xlsx",
                   help="출력 서식(템플릿) 파일")
    p.add_argument("--out", default="최종평가위원 섭외우선(안)_결과.xlsx",
                   help="결과 저장 파일")
    p.add_argument("--panel", default=None,
                   help="A열(구분/패널명). 미지정 시 템플릿 기존 값 유지")
    a = p.parse_args(argv)

    rows, n = build(a.seon, a.dan, a.template, a.out, a.panel)

    # 요약 출력
    from collections import Counter
    cnt = Counter(r["history"] for r in rows)
    print(f"[완료] 총 {n}행 -> {a.out}")
    for label in sorted(cnt, key=lambda x: PRIORITY.get(x, 99)):
        print(f"  - {label}: {cnt[label]}명")
    return 0


if __name__ == "__main__":
    sys.exit(main())
